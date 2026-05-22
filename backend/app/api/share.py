import os
import json
from fastapi import APIRouter, Depends, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from typing import Optional
from app.services.account_manager import account_manager
from app.core.config import settings
from app.core.database import async_session
from app.models.schema import ShareAnalysisResult, ShareAnalysisState, P115Account, SharePushTask
from sqlalchemy import select, delete, desc, or_, func
from loguru import logger
import asyncio
import time
from datetime import datetime
import io

router = APIRouter(prefix="/share", tags=["share"])


def format_size(size_bytes):
    if not size_bytes or size_bytes == 0:
        return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    import math
    try:
        i = int(math.floor(math.log(size_bytes, 1024)))
        p = math.pow(1024, i)
        s = round(size_bytes / p, 2)
        return "%s %s" % (s, size_name[i])
    except:
        return "0B"


def get_share_status(item: dict) -> dict:
    state = item.get("share_state")
    have_vio = item.get("have_vio_file", 0)

    is_violated = False
    is_expired = False
    is_reviewing = False
    is_invalid = False

    if state == 6:
        # share_state==6：链接已被 115 因违规撤销，公开访问返回 errno 4100009「链接已失效」
        status_text = "已失效"
        is_violated = True
        is_invalid = True
    elif have_vio == 1:
        # have_vio_file==1：分享包含违规内容标记，但链接本身仍可访问（页面显示「文件内含违规内容」警告）
        status_text = "违规"
        is_violated = True
    elif state == 7:
        status_text = "已过期"
        is_expired = True
    elif state == 4:
        status_text = "审核中"
        is_reviewing = True
    else:
        status_text = "正常"

    return {
        "status_text": status_text,
        "is_violated": is_violated,
        "is_invalid": is_invalid,
        "is_expired": is_expired,
        "is_reviewing": is_reviewing,
    }


# ──── 内存分析状态（按 account_id 分组）─────────────────────────────────
# account_id(int) -> state dict
_analysis_states: dict[int, dict] = {}

# ──── 批量取消状态（按 account_id 分组）─────────────────────────────────
_cancel_states: dict[int, dict] = {}

def _get_cancel_state(account_id: int) -> dict:
    if account_id not in _cancel_states:
        _cancel_states[account_id] = {
            "is_canceling": False,
            "total": 0,
            "done": 0,
            "failed": 0,
        }
    return _cancel_states[account_id]

def _get_state(account_id: int) -> dict:
    """获取或初始化某账号的内存分析状态"""
    if account_id not in _analysis_states:
        _analysis_states[account_id] = {
            "is_analyzing": False,
            "total": 0,
            "normal": 0,
            "violated": 0,
            "expired": 0,
            "reviewing": 0,
            "scanned": 0,
            "last_updated": None,
        }
    return _analysis_states[account_id]


async def _load_state_from_db(account_id: int):
    """从数据库加载账号的分析状态到内存"""
    async with async_session() as session:
        result = await session.execute(
            select(ShareAnalysisState).where(ShareAnalysisState.account_id == account_id)
        )
        row = result.scalar_one_or_none()
        if row:
            state = _get_state(account_id)
            state.update({
                "is_analyzing": False,
                "total": row.total,
                "normal": row.normal,
                "violated": row.violated,
                "expired": row.expired,
                "reviewing": row.reviewing,
                "scanned": row.scanned,
                "last_updated": row.last_updated,
            })


async def _save_state_to_db(account_id: int):
    """将内存分析状态持久化到数据库"""
    state = _get_state(account_id)
    async with async_session() as session:
        result = await session.execute(
            select(ShareAnalysisState).where(ShareAnalysisState.account_id == account_id)
        )
        row = result.scalar_one_or_none()
        if row:
            row.is_analyzing = state["is_analyzing"]
            row.total = state["total"]
            row.normal = state["normal"]
            row.violated = state["violated"]
            row.expired = state["expired"]
            row.reviewing = state["reviewing"]
            row.scanned = state["scanned"]
            row.last_updated = state["last_updated"]
        else:
            row = ShareAnalysisState(
                account_id=account_id,
                is_analyzing=state["is_analyzing"],
                total=state["total"],
                normal=state["normal"],
                violated=state["violated"],
                expired=state["expired"],
                reviewing=state["reviewing"],
                scanned=state["scanned"],
                last_updated=state["last_updated"],
            )
            session.add(row)
        await session.commit()


async def perform_share_analysis(account_id: int):
    """对指定账号执行全量分享分析"""
    svc = account_manager.get_service(account_id)
    if not svc or not svc.client:
        logger.warning(f"账号 {account_id} 未连接，无法执行分析")
        state = _get_state(account_id)
        state["is_analyzing"] = False
        return

    state = _get_state(account_id)
    state["is_analyzing"] = True
    state["normal"] = 0
    state["violated"] = 0
    state["expired"] = 0
    state["reviewing"] = 0
    state["scanned"] = 0

    # 清空该账号的旧结果
    async with async_session() as session:
        await session.execute(
            delete(ShareAnalysisResult).where(ShareAnalysisResult.account_id == account_id)
        )
        await session.commit()

    try:
        offset = 0
        limit = 100

        while True:
            payload = {
                "offset": offset,
                "limit": limit,
                "order": "create_time",
                "asc": 0,
                "show_cancel_share": 0
            }

            resp = await svc.client.share_list(payload, async_=True)
            if not resp.get("state"):
                break

            items = resp.get("list", [])
            if not items:
                break

            state["total"] = resp.get("count", 0)

            rows_to_insert = []
            for item in items:
                status_info = get_share_status(item)

                if status_info["is_violated"]:
                    state["violated"] += 1
                elif status_info["is_expired"]:
                    state["expired"] += 1
                elif status_info["is_reviewing"]:
                    state["reviewing"] += 1
                else:
                    state["normal"] += 1

                state["scanned"] += 1

                create_time = item.get("create_time", 0)
                create_time_str = datetime.fromtimestamp(create_time).strftime('%Y-%m-%d %H:%M:%S') if create_time else "-"

                rows_to_insert.append(ShareAnalysisResult(
                    account_id=account_id,
                    snap_id=item.get("snap_id"),
                    share_code=item.get("share_code"),
                    share_title=item.get("share_title"),
                    share_url=item.get("share_url"),
                    receive_code=item.get("receive_code", ""),
                    file_size=item.get("file_size", 0),
                    size_text=format_size(item.get("file_size", 0)),
                    create_time=create_time_str,
                    create_timestamp=create_time,
                    share_state=item.get("share_state"),
                    status_text=status_info["status_text"],
                    is_violated=status_info["is_violated"],
                    is_invalid=status_info["is_invalid"],
                    is_expired=status_info["is_expired"],
                    is_reviewing=status_info["is_reviewing"],
                    receive_count=item.get("receive_count", 0),
                ))

            # 批量写入数据库
            async with async_session() as session:
                session.add_all(rows_to_insert)
                await session.commit()

            # 同步状态到数据库
            await _save_state_to_db(account_id)

            if len(items) < limit:
                break

            offset += limit
            await asyncio.sleep(1.2)

        state["last_updated"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        await _save_state_to_db(account_id)

    except Exception as e:
        logger.error(f"账号 {account_id} 分析失败: {e}")
    finally:
        state["is_analyzing"] = False
        await _save_state_to_db(account_id)


async def _refresh_analysis_stats(account_id: int):
    """从数据库重新计算分析统计并更新内存状态及数据库"""
    async with async_session() as session:
        rows = (await session.execute(
            select(ShareAnalysisResult).where(ShareAnalysisResult.account_id == account_id)
        )).scalars().all()
    state = _get_state(account_id)
    state["total"] = len(rows)
    state["normal"] = sum(1 for r in rows if not r.is_violated and not r.is_expired and not r.is_reviewing)
    state["violated"] = sum(1 for r in rows if r.is_violated)
    state["expired"] = sum(1 for r in rows if r.is_expired)
    state["reviewing"] = sum(1 for r in rows if r.is_reviewing)
    state["scanned"] = state["total"]
    await _save_state_to_db(account_id)


async def perform_batch_cancel(account_id: int):
    """批量取消已失效和已过期的分享"""
    state = _get_cancel_state(account_id)
    state["is_canceling"] = True
    state["done"] = 0
    state["failed"] = 0

    svc = account_manager.get_service(account_id)
    if not svc or not svc.client:
        state["is_canceling"] = False
        return

    try:
        async with async_session() as session:
            stmt = select(ShareAnalysisResult).where(
                ShareAnalysisResult.account_id == account_id,
                or_(ShareAnalysisResult.is_invalid == True, ShareAnalysisResult.is_expired == True)
            )
            rows = (await session.execute(stmt)).scalars().all()
            targets = [(r.id, r.share_code) for r in rows if r.share_code]

        state["total"] = len(targets)
        canceled_ids = []

        for i, (db_id, share_code) in enumerate(targets):
            try:
                resp = await svc.client.share_update(
                    {"share_code": share_code, "action": "cancel"},
                    async_=True
                )
                if resp.get("state"):
                    canceled_ids.append(db_id)
                    state["done"] += 1
                else:
                    state["failed"] += 1
                    logger.warning(f"取消分享 {share_code} 返回失败: {resp.get('error', '')}")
            except Exception as e:
                state["failed"] += 1
                logger.warning(f"取消分享 {share_code} 异常: {e}")

            if (i + 1) % 10 == 0:
                await asyncio.sleep(0.5)

        if canceled_ids:
            async with async_session() as session:
                await session.execute(
                    delete(ShareAnalysisResult).where(ShareAnalysisResult.id.in_(canceled_ids))
                )
                await session.commit()

        await _refresh_analysis_stats(account_id)
        logger.info(f"批量取消完成: 成功 {state['done']}，失败 {state['failed']}")

    except Exception as e:
        logger.error(f"批量取消分享失败: {e}")
    finally:
        state["is_canceling"] = False


    async with async_session() as session:
        result = await session.execute(
            select(P115Account).where(P115Account.enabled == True).order_by(P115Account.priority).limit(1)
        )
        acc = result.scalar_one_or_none()
        return acc.id if acc else None


@router.post("/analyze")
async def start_analysis(background_tasks: BackgroundTasks, account_id: Optional[int] = None):
    """开始全量分享分析"""
    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"state": False, "message": "未配置任何账号"}

    state = _get_state(account_id)
    if state["is_analyzing"]:
        return {"state": True, "message": "该账号分析正在进行中"}

    background_tasks.add_task(perform_share_analysis, account_id)
    return {"state": True, "message": "分析已启动", "account_id": account_id}


@router.post("/reset")
async def reset_analysis(account_id: Optional[int] = None):
    """重置指定账号的分析结果"""
    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"state": False, "message": "未配置任何账号"}

    # 重置内存状态
    _analysis_states[account_id] = {
        "is_analyzing": False,
        "total": 0, "normal": 0, "violated": 0,
        "expired": 0, "reviewing": 0, "scanned": 0,
        "last_updated": None,
    }

    # 清空数据库
    async with async_session() as session:
        await session.execute(
            delete(ShareAnalysisResult).where(ShareAnalysisResult.account_id == account_id)
        )
        await session.execute(
            delete(ShareAnalysisState).where(ShareAnalysisState.account_id == account_id)
        )
        await session.commit()

    return {"state": True, "message": "分析数据已重置"}


@router.get("/analysis-status")
async def get_analysis_status(account_id: Optional[int] = None):
    """获取指定账号的分析状态"""
    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"account_id": None, "is_analyzing": False, "total": 0,
                "normal": 0, "violated": 0, "expired": 0, "reviewing": 0,
                "scanned": 0, "last_updated": None}

    # 若内存中没有状态，先从 DB 加载
    if account_id not in _analysis_states:
        await _load_state_from_db(account_id)

    state = _get_state(account_id).copy()
    state["account_id"] = account_id
    return state


@router.get("/list")
async def list_shares(
    offset: int = 0,
    limit: int = 32,
    order: str = "create_time",
    asc: int = 0,
    search_value: Optional[str] = None,
    status_filter: str = "all",
    account_id: Optional[int] = None,
):
    """从数据库中获取分享列表（按账号过滤）"""
    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"state": True, "count": 0, "list": []}

    async with async_session() as session:
        stmt = select(ShareAnalysisResult).where(ShareAnalysisResult.account_id == account_id)

        # 状态过滤
        if status_filter == "normal":
            stmt = stmt.where(
                ShareAnalysisResult.is_violated == False,
                ShareAnalysisResult.is_expired == False,
                ShareAnalysisResult.is_reviewing == False,
            )
        elif status_filter == "violated":
            stmt = stmt.where(ShareAnalysisResult.is_violated == True)
        elif status_filter == "expired":
            stmt = stmt.where(ShareAnalysisResult.is_expired == True)
        elif status_filter == "reviewing":
            stmt = stmt.where(ShareAnalysisResult.is_reviewing == True)

        # 搜索
        if search_value:
            stmt = stmt.where(ShareAnalysisResult.share_title.ilike(f"%{search_value}%"))

        # 排序
        from sqlalchemy import asc as sa_asc, desc as sa_desc
        sort_col_map = {
            "create_time": ShareAnalysisResult.create_timestamp,
            "file_size": ShareAnalysisResult.file_size,
            "receive_count": ShareAnalysisResult.receive_count,
            "share_title": ShareAnalysisResult.share_title,
            "share_state": ShareAnalysisResult.share_state,
        }
        sort_col = sort_col_map.get(order, ShareAnalysisResult.create_timestamp)
        stmt = stmt.order_by(sa_asc(sort_col) if asc == 1 else sa_desc(sort_col))

        # 总数（不分页）
        from sqlalchemy import func
        count_stmt = select(func.count()).select_from(stmt.subquery())
        total_count = (await session.execute(count_stmt)).scalar_one()

        # 分页
        stmt = stmt.offset(offset).limit(limit)
        rows = (await session.execute(stmt)).scalars().all()

    results = [
        {
            "id": r.id,
            "snap_id": r.snap_id,
            "share_code": r.share_code,
            "share_title": r.share_title,
            "share_url": r.share_url,
            "receive_code": r.receive_code,
            "file_size": r.file_size,
            "size_text": r.size_text,
            "create_time": r.create_time,
            "create_timestamp": r.create_timestamp,
            "share_state": r.share_state,
            "status_text": r.status_text,
            "is_violated": r.is_violated,
            "is_invalid": r.is_invalid,
            "is_expired": r.is_expired,
            "is_reviewing": r.is_reviewing,
            "receive_count": r.receive_count,
        }
        for r in rows
    ]

    return {"state": True, "count": total_count, "list": results}


async def _get_filtered_results(account_id: int, status_filter: str, search_value: Optional[str]) -> list:
    """从数据库取过滤后的结果列表（用于导出）"""
    async with async_session() as session:
        stmt = select(ShareAnalysisResult).where(ShareAnalysisResult.account_id == account_id)
        if status_filter == "normal":
            stmt = stmt.where(
                ShareAnalysisResult.is_violated == False,
                ShareAnalysisResult.is_expired == False,
                ShareAnalysisResult.is_reviewing == False,
            )
        elif status_filter == "violated":
            stmt = stmt.where(ShareAnalysisResult.is_violated == True)
        elif status_filter == "expired":
            stmt = stmt.where(ShareAnalysisResult.is_expired == True)
        elif status_filter == "reviewing":
            stmt = stmt.where(ShareAnalysisResult.is_reviewing == True)
        if search_value:
            stmt = stmt.where(ShareAnalysisResult.share_title.ilike(f"%{search_value}%"))
        rows = (await session.execute(stmt)).scalars().all()
    return rows


@router.get("/export/json")
async def export_json(
    search_value: Optional[str] = None,
    status_filter: str = "all",
    account_id: Optional[int] = None,
):
    """导出分析结果为 JSON 文件"""
    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"state": False, "error": "未配置任何账号"}

    rows = await _get_filtered_results(account_id, status_filter, search_value)
    if not rows:
        return {"state": False, "error": "没有可导出的数据"}

    results = [
        {
            "snap_id": r.snap_id, "share_code": r.share_code,
            "share_title": r.share_title, "share_url": r.share_url,
            "receive_code": r.receive_code, "file_size": r.file_size,
            "size_text": r.size_text, "create_time": r.create_time,
            "create_timestamp": r.create_timestamp, "share_state": r.share_state,
            "status_text": r.status_text, "is_violated": r.is_violated,
            "is_expired": r.is_expired, "is_reviewing": r.is_reviewing,
            "receive_count": r.receive_count,
        }
        for r in rows
    ]

    export_data = {
        "export_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "account_id": account_id,
        "total_count": len(results),
        "data": results,
    }
    json_str = json.dumps(export_data, ensure_ascii=False, indent=2)
    return StreamingResponse(
        io.BytesIO(json_str.encode('utf-8')),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=share_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        }
    )


@router.get("/export/excel")
async def export_excel(
    search_value: Optional[str] = None,
    status_filter: str = "all",
    account_id: Optional[int] = None,
):
    """导出分析结果为 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return {"state": False, "error": "缺少 openpyxl 依赖，请安装: pip install openpyxl"}

    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"state": False, "error": "未配置任何账号"}

    rows = await _get_filtered_results(account_id, status_filter, search_value)
    if not rows:
        return {"state": False, "error": "没有可导出的数据"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分享分析结果"

    headers = ["分享名称", "状态", "分享链接", "提取码", "文件大小", "创建时间", "接收次数", "分享码"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([
            r.share_title or "",
            r.status_text or "",
            r.share_url or "",
            r.receive_code or "",
            r.size_text or "",
            r.create_time or "",
            r.receive_count or 0,
            r.share_code or "",
        ])

    column_widths = [40, 12, 50, 12, 15, 20, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=share_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )


@router.post("/push-to-channel")
async def push_to_channel(
    background_tasks: BackgroundTasks,
    channel_id: str = Query(..., description="目标频道ID"),
    channel_name: Optional[str] = Query(None, description="频道名称"),
    share_ids: Optional[str] = Query(None, description="分享ID列表，逗号分隔"),
    push_all: bool = Query(False, description="是否推送全部"),
    account_id: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    interval_min: int = Query(3, description="推送间隔(秒)下限"),
    interval_max: int = Query(5, description="推送间隔(秒)上限"),
    skip_count: int = Query(0, ge=0, description="跳过前N条"),
    stop_at: int = Query(0, ge=0, description="到第M条停止，0表示不限制"),
):
    """创建推送任务"""
    from app.services.tg_bot import tg_service

    if not tg_service.bot or not tg_service.is_connected:
        return {"state": False, "error": "Telegram Bot 未连接"}

    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"state": False, "error": "未配置任何账号"}

    if stop_at > 0 and stop_at <= skip_count:
        return {"state": False, "error": "第几条停止必须大于跳过前几条（0 表示不限制）"}

    # 解析分享ID列表
    id_list = []
    if share_ids:
        try:
            id_list = [int(x.strip()) for x in share_ids.split(',') if x.strip()]
        except:
            return {"state": False, "error": "分享ID格式错误"}

    # 获取要推送的分享链接
    async with async_session() as session:
        stmt = select(ShareAnalysisResult).where(
            ShareAnalysisResult.account_id == account_id,
            ShareAnalysisResult.is_expired == False,
        )

        # 时间范围过滤
        if start_date:
            try:
                start_ts = int(datetime.strptime(start_date, '%Y-%m-%d').timestamp())
                stmt = stmt.where(ShareAnalysisResult.create_timestamp >= start_ts)
            except:
                pass

        if end_date:
            try:
                end_ts = int(datetime.strptime(end_date, '%Y-%m-%d').timestamp())
                stmt = stmt.where(ShareAnalysisResult.create_timestamp <= end_ts)
            except:
                pass

        # 选择特定ID或全部
        if not push_all and id_list:
            stmt = stmt.where(ShareAnalysisResult.id.in_(id_list))

        # 固定排序，确保 skip/stop 语义稳定
        stmt = stmt.order_by(desc(ShareAnalysisResult.create_timestamp), desc(ShareAnalysisResult.id))

        rows = (await session.execute(stmt)).scalars().all()

        # 对过滤结果执行范围内跳过/停止
        end_index = stop_at if stop_at > 0 else None
        rows = rows[skip_count:end_index]

        if not rows:
            return {"state": False, "error": "没有符合条件的分享链接"}

        # 创建推送任务
        task = SharePushTask(
            account_id=account_id,
            channel_id=channel_id,
            channel_name=channel_name or channel_id,
            status="running",
            total_count=len(rows),
            share_ids=[r.id for r in rows],
            failed_ids=[],
            interval_min=interval_min,
            interval_max=interval_max,
        )
        session.add(task)
        await session.commit()
        await session.refresh(task)
        task_id = task.id

    # 启动后台推送
    background_tasks.add_task(perform_push_task, task_id)

    return {
        "state": True,
        "message": f"推送任务已创建，共 {len(rows)} 条分享链接",
        "task_id": task_id,
        "total": len(rows)
    }


async def perform_push_task(task_id: int):
    """执行推送任务"""
    from app.services.tg_bot import tg_service

    async with async_session() as session:
        result = await session.execute(select(SharePushTask).where(SharePushTask.id == task_id))
        task = result.scalar_one_or_none()
        if not task:
            return

        share_ids = task.share_ids
        channel_id = task.channel_id

        for idx, share_id in enumerate(share_ids):
            # 检查任务状态
            await session.refresh(task)
            if task.status == "cancelled":
                logger.info(f"推送任务 {task_id} 已取消")
                break
            elif task.status == "paused":
                logger.info(f"推送任务 {task_id} 已暂停")
                while task.status == "paused":
                    await asyncio.sleep(2)
                    await session.refresh(task)
                    if task.status == "cancelled":
                        break
            
            if task.status == "cancelled":
                break

            # 获取分享信息
            result = await session.execute(
                select(ShareAnalysisResult).where(ShareAnalysisResult.id == share_id)
            )
            row = result.scalar_one_or_none()
            if not row:
                task.fail_count += 1
                continue

            try:
                # 构建消息
                msg_text = f"📦 {row.share_title}\n\n"

                # 链接：如果有提取码，直接拼接到URL
                share_url = row.share_url
                if row.receive_code:
                    share_url = f"{row.share_url}?password={row.receive_code}"
                msg_text += f"🔗 链接: {share_url}\n"

                msg_text += f"📊 大小: {row.size_text}\n"
                msg_text += f"📅 分享时间: {row.create_time}\n"
                msg_text += f"👥 接收次数: {row.receive_count}"

                await tg_service.bot.send_message(
                    chat_id=channel_id,
                    text=msg_text,
                )
                task.success_count += 1
                logger.info(f"推送成功: {row.share_title}")
            except Exception as e:
                logger.error(f"推送失败 {share_id}: {e}")
                task.fail_count += 1
                # 记录失败的 ID
                if task.failed_ids is None:
                    task.failed_ids = []
                # 避免重复记录
                if share_id not in task.failed_ids:
                    task.failed_ids = task.failed_ids + [share_id] # SQLAlchemy JSON mutation check

            task.current_index = idx + 1
            await session.commit()
            import random
            delay = random.uniform(task.interval_min, task.interval_max)
            await asyncio.sleep(delay)  # 增加延迟以符合 Telegram 限制 (每分钟最多推送20条)

        # 任务完成
        if task.status != "cancelled":
            task.status = "completed"
        await session.commit()
        logger.info(f"推送任务 {task_id} 完成: 成功 {task.success_count}, 失败 {task.fail_count}")


@router.get("/push-tasks")
async def get_push_tasks(account_id: Optional[int] = None):
    """获取推送任务列表"""
    if account_id is None:
        account_id = await _get_default_account_id()

    async with async_session() as session:
        stmt = select(SharePushTask).where(SharePushTask.account_id == account_id).order_by(SharePushTask.created_at.desc())
        rows = (await session.execute(stmt)).scalars().all()

    return {
        "state": True,
        "tasks": [
            {
                "id": t.id,
                "channel_id": t.channel_id,
                "channel_name": t.channel_name,
                "status": t.status,
                "total_count": t.total_count,
                "success_count": t.success_count,
                "fail_count": t.fail_count,
                "current_index": t.current_index,
                "created_at": t.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            }
            for t in rows
        ]
    }


@router.post("/push-task/{task_id}/pause")
async def pause_push_task(task_id: int):
    """暂停推送任务"""
    async with async_session() as session:
        result = await session.execute(select(SharePushTask).where(SharePushTask.id == task_id))
        task = result.scalar_one_or_none()
        if not task:
            return {"state": False, "error": "任务不存在"}

        if task.status != "running":
            return {"state": False, "error": "任务未在运行中"}

        task.status = "paused"
        await session.commit()

    return {"state": True, "message": "任务已暂停"}


@router.post("/push-task/{task_id}/resume")
async def resume_push_task(task_id: int):
    """恢复推送任务"""
    async with async_session() as session:
        result = await session.execute(select(SharePushTask).where(SharePushTask.id == task_id))
        task = result.scalar_one_or_none()
        if not task:
            return {"state": False, "error": "任务不存在"}

        if task.status != "paused":
            return {"state": False, "error": "任务未暂停"}

        task.status = "running"
        await session.commit()

    return {"state": True, "message": "任务已恢复"}


@router.post("/push-task/{task_id}/cancel")
async def cancel_push_task(task_id: int):
    """取消推送任务"""
    async with async_session() as session:
        result = await session.execute(select(SharePushTask).where(SharePushTask.id == task_id))
        task = result.scalar_one_or_none()
        if not task:
            return {"state": False, "error": "任务不存在"}

        if task.status in ["completed", "cancelled"]:
            return {"state": False, "error": "任务已结束"}

        task.status = "cancelled"
        await session.commit()

    return {"state": True, "message": "任务已取消"}


@router.post("/push-task/{task_id}/retry")
async def retry_push_task(background_tasks: BackgroundTasks, task_id: int):
    """重试推送任务中的失败项"""
    async with async_session() as session:
        result = await session.execute(select(SharePushTask).where(SharePushTask.id == task_id))
        old_task = result.scalar_one_or_none()
        if not old_task:
            return {"state": False, "error": "原任务不存在"}

        if not old_task.failed_ids:
            return {"state": False, "error": "没有失败的项目可以重试"}

        failed_ids = old_task.failed_ids
        
        # 创建新任务
        new_task = SharePushTask(
            account_id=old_task.account_id,
            channel_id=old_task.channel_id,
            channel_name=f"{old_task.channel_name} (重试)",
            status="running",
            total_count=len(failed_ids),
            share_ids=failed_ids,
            failed_ids=[],
        )
        session.add(new_task)
        await session.commit()
        await session.refresh(new_task)
        new_task_id = new_task.id

    # 启动后台推送
    background_tasks.add_task(perform_push_task, new_task_id)

    return {
        "state": True,
        "message": f"已创建重试任务，共 {len(failed_ids)} 条",
        "task_id": new_task_id
    }


@router.delete("/push-task/{task_id}")
async def delete_push_task(task_id: int):
    """删除推送任务"""
    async with async_session() as session:
        await session.execute(delete(SharePushTask).where(SharePushTask.id == task_id))
        await session.commit()

    return {"state": True, "message": "任务已删除"}


# ──── 取消分享 ─────────────────────────────────────────────────────────────────

@router.post("/cancel")
async def cancel_share(share_code: str, account_id: Optional[int] = None):
    """取消单个分享链接"""
    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"state": False, "error": "未配置任何账号"}

    svc = account_manager.get_service(account_id)
    if not svc or not svc.client:
        return {"state": False, "error": "账号未连接"}

    try:
        resp = await svc.client.share_update(
            {"share_code": share_code, "action": "cancel"},
            async_=True
        )
        if not resp.get("state"):
            return {"state": False, "error": resp.get("error", "取消失败")}

        async with async_session() as session:
            await session.execute(
                delete(ShareAnalysisResult).where(
                    ShareAnalysisResult.account_id == account_id,
                    ShareAnalysisResult.share_code == share_code,
                )
            )
            await session.commit()

        await _refresh_analysis_stats(account_id)
        return {"state": True, "message": "取消成功"}
    except Exception as e:
        logger.error(f"取消分享 {share_code} 失败: {e}")
        return {"state": False, "error": str(e)}


@router.post("/cancel-invalid-expired")
async def cancel_invalid_expired(
    background_tasks: BackgroundTasks,
    account_id: Optional[int] = None,
):
    """批量取消所有已失效和已过期的分享"""
    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"state": False, "error": "未配置任何账号"}

    state = _get_cancel_state(account_id)
    if state["is_canceling"]:
        return {"state": True, "message": "正在取消中，请稍候"}

    async with async_session() as session:
        count = (await session.execute(
            select(func.count()).select_from(
                select(ShareAnalysisResult).where(
                    ShareAnalysisResult.account_id == account_id,
                    or_(ShareAnalysisResult.is_invalid == True, ShareAnalysisResult.is_expired == True)
                ).subquery()
            )
        )).scalar_one()

    if count == 0:
        return {"state": True, "message": "没有需要取消的分享", "count": 0}

    background_tasks.add_task(perform_batch_cancel, account_id)
    return {"state": True, "message": f"已启动批量取消任务，共 {count} 条", "count": count}


@router.get("/cancel-status")
async def get_cancel_status(account_id: Optional[int] = None):
    """获取批量取消任务状态"""
    if account_id is None:
        account_id = await _get_default_account_id()
    if account_id is None:
        return {"is_canceling": False, "total": 0, "done": 0, "failed": 0}

    state = _get_cancel_state(account_id)
    return {
        "account_id": account_id,
        "is_canceling": state["is_canceling"],
        "total": state["total"],
        "done": state["done"],
        "failed": state["failed"],
    }
